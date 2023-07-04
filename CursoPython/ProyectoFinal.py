from tkinter import *
import tkinter as tk
from tkinter import messagebox as mb
from openpyxl import *
import serial
import os
import threading
from PIL import Image, ImageTk

RUTA = r"C:\Users\omarr\OneDrive\Documentos\CursoPython\datos.xlsx"
PUERTO = "COM5"
BG = "#4e6d9b"
FG = "#ffffff"
FGON = "#408f4b"
FGOFF = "#ff1111"
FONTC = ("Arial", 15)


""" 
        POR:                                                                         OMAR RAMIREZ HERNANDEZ
        
        
        LETRAS DE ARDUINO : 
        
        LED
        	o : Encender
        	a : apagar


        Alarma
        	e : encender
        	f : apagar


        Ultrasonido
        	u : recivir


        Puerta
        	p : encender
        	q : apagar


        Gas
        	g : recivir


        Humedad
        	h : recivir
         
        Llave (simulada con un led)
          l : encender
          i : apagar
"""


try:
    arduino = serial.Serial(PUERTO, 9600)
except:
    print("Conexion fallida")


def closeW():
    arduino.close()
    root.destroy()


def readSaveUltrasonic():
    try:
        arduino.write("u\n").encode()
    except:
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")
        print("No arduino")
    libro = load_workbook(RUTA)
    hoja2 = libro.active
    hoja2.title = "hoja2"
    try:
        while arduino.is_open:
            datos = arduino.readline().decode().strip()
            etiqueta4.config(text=datos)
            next_row = hoja2.max_row + 1
            hoja2.cell(row=next_row, column=1).value = datos
            libro.save(RUTA)
    except:
        print("No arduino")


def initThreadUltrasonic():
    hilo = threading.Thread(target=readSaveUltrasonic)
    hilo.daemon = True
    hilo.start()


def readSaveGas():
    try:
        arduino.write("g\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")
    libro = load_workbook(RUTA)
    hoja2 = libro.active
    hoja2.title = "hoja2"
    try:
        while arduino.is_open:
            datos = arduino.readline().decode().strip()
            etiqueta5.config(text=datos)
            next_row = hoja2.max_row + 1
            hoja2.cell(row=next_row, column=2).value = datos
            libro.save(RUTA)
    except:
        print("No arduino")


def initThreadGas():
    hilo = threading.Thread(target=readSaveGas)
    hilo.daemon = True
    hilo.start()


def readSaveHumedad():
    try:
        arduino.write("h\n").encode()
    except:
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")
        print("No arduino")
    libro = load_workbook(RUTA)
    hoja2 = libro.active
    hoja2.title = "hoja2"
    try:
        while arduino.is_open:
            datos = arduino.readline().decode().strip()
            etiqueta6.config(text=datos)
            next_row = hoja2.max_row + 1
            hoja2.cell(row=next_row, column=3).value = datos
            libro.save(RUTA)
    except:
        print("No arduino")


def initThreadHumedad():
    hilo = threading.Thread(target=readSaveHumedad)
    hilo.daemon = True
    hilo.start()


def on1():
    try:
        arduino.write("o\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")
    etiqueta1.config(text="Led Encendido", fg=FGON)


def off1():
    try:
        arduino.write("a\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")
    etiqueta1.config(text="Led Apagado", fg=FGOFF)


def onAlarm():
    etiqueta2.config(text="Alarma encendida", fg=FGON)
    try:
        arduino.write("e\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")


def offAlarm():
    etiqueta2.config(text="Alarma apagada", fg=FGOFF)
    try:
        arduino.write("f\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")


def onDoor():
    etiqueta3.config(text="Puerta abierta", fg=FGON)
    try:
        arduino.write("p\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")


def offDoor():
    etiqueta3.config(text="Puerta cerrada", fg=FGOFF)
    try:
        arduino.write("q\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")


def onLlave():
    etiqueta7.config(text="Llave abierta", fg=FGON)
    try:
        arduino.write("l\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")


def offLlave():
    etiqueta7.config(text="Llave cerrada", fg=FGOFF)
    try:
        arduino.write("i\n").encode()
    except:
        print("No arduino")
        mb.showerror("Smart Home | Error", "Arduino no esta conectado")


def closeW():
    # arduino.close()
    root.destroy()


if not os.path.exists(RUTA):
    libro = Workbook()
    libro.save(RUTA)

root = Tk()
root.title("Smart Home")
root.geometry("820x700")
# root.config(background="#4e6d9b")
# root.config(background="#639c3f")
root.protocol("W_DELETE_WINDOW", closeW)
root.config(background="#dddddd")

title = Label(root, text="Smart Home", font=("Arial", 40))
title.grid(row=0, column=0, padx=30, pady=10)

image = Image.open("R.png")
image = image.resize((120, 120))
image = ImageTk.PhotoImage(image)
icon = Label(root, image=image, width=130, height=130)
icon.grid(row=0, column=2, padx=10, pady=10)

btn_on1 = Button(
    root,
    text="Encender Led",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=34,
)
btn_on1.grid(row=1, column=0)
btn_on1[COMMAND] = on1
btn_off1 = Button(
    root,
    text="Apagar Led",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=15,
)
btn_off1.grid(row=1, column=1)
btn_off1[COMMAND] = off1
etiqueta1 = Label(
    root,
    text="Led Apagado",
    fg=FGOFF,
    padx=40,
    pady=10,
    font=FONTC,
    width=10,
)
etiqueta1.grid(row=1, column=2)


btn_on2 = Button(
    root,
    text="Encender Alarma",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=34,
    command=onAlarm,
)
btn_on2.grid(row=2, column=0)
btn_off2 = Button(
    root,
    text="Apagar Alarma",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=15,
    command=offAlarm,
)
btn_off2.grid(row=2, column=1)
etiqueta2 = Label(
    root,
    text="Alarma Apagada",
    fg=FGOFF,
    padx=40,
    pady=10,
    font=FONTC,
    width=10,
)
etiqueta2.grid(row=2, column=2)


# Para cerrar la puerta se usará un servo motor
btn_on3 = Button(
    root,
    text="Abrir puerta",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=34,
    command=onDoor,
)
btn_on3.grid(row=3, column=0)
btn_off3 = Button(
    root,
    text="Cerrar puerta",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=15,
    command=offDoor,
)
btn_off3.grid(row=3, column=1)
etiqueta3 = Label(
    root,
    text="Puerta cerrada",
    fg=FGOFF,
    padx=40,
    pady=10,
    font=FONTC,
    width=10,
)
etiqueta3.grid(row=3, column=2)


# Para saber el nivel del agua se utilizara un sensor Ultrasonico
btn_on4 = Button(
    root,
    text="Consultar nivel del tanque de agua",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=34,
    command=initThreadUltrasonic,
)
btn_on4.grid(
    row=4,
    column=0,
    columnspan=2,
)
etiqueta4 = Label(
    root,
    text="0%",
    fg=BG,
    padx=40,
    pady=10,
    font=FONTC,
    width=10,
)
etiqueta4.grid(row=4, column=2)


# Para consultar la contaminacion del aire se usará un sensor de gas
btn_on5 = Button(
    root,
    text="Consultar Toxisidad del aire",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=34,
    command=initThreadGas,
)
btn_on5.grid(
    row=5,
    column=0,
    columnspan=2,
)
etiqueta5 = Label(
    root,
    text="Nivel: bajo",
    fg=FGOFF,
    padx=40,
    pady=10,
    font=FONTC,
    width=10,
)
etiqueta5.grid(row=5, column=2)


btn_on6 = Button(
    root,
    text="Consultar humedad de la planta",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=34,
    command=initThreadHumedad,
)
btn_on6.grid(
    row=6,
    column=0,
    columnspan=2,
)
etiqueta6 = Label(
    root,
    text="Numero: ",
    fg=BG,
    padx=40,
    pady=10,
    font=FONTC,
    width=10,
)
etiqueta6.grid(row=6, column=2)

btn_on7 = Button(
    root,
    text="Abrir llave(Agua)",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=34,
    command=onLlave,
)
btn_on7.grid(
    row=7,
    column=0,
)
btn_off7 = Button(
    root,
    text="Cerrar llave (Agua)",
    padx=10,
    pady=10,
    background=BG,
    fg=FG,
    font=FONTC,
    width=15,
    command=offLlave,
)
btn_off7.grid(
    row=7,
    column=1,
)
etiqueta7 = Label(
    root,
    text="Llave cerrada",
    fg=FGOFF,
    padx=40,
    pady=10,
    font=FONTC,
    width=10,
)
etiqueta7.grid(
    row=7,
    column=2,
)

nameProgramer = Label(
    root,
    text="Programado por: Omar Ramirez Hernandez...",
    fg="#639c3f",
    padx=10,
    pady=10,
    font=("Arial", 25),
)
nameProgramer.grid(row=8, column=0, columnspan=3)
root.eval("tk::PlaceWindow . center")
root.update()
root.mainloop()
