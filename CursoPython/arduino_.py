from tkinter import*
import serial
import os
import threading
from openpyxl import *

carpeta = r"C:\Users\xedkr\Desktop\pracExc1"
ruta = r"C:\Users\xedkr\Desktop\pracExc1\datos.xlsx"
puerto = "COM5"

# Crear directorio si no existe
if not os.path.exists(carpeta):
    os.makedirs(carpeta)

# Crear archivo de Excel
if not os.path.exists(ruta):
    libro = Workbook()
    libro.save(ruta)

# Crear puerto serial
try:
    arduino = serial.Serial(puerto, 9600)
except:
    print("No se estableció la conexión")


#para cerrar ventana
def cerrar_ventana():
    arduino.close()
    root.destroy()


# Función para leer y guardar datos en archivo de Excel
def read_and_save():
    # Cargar el archivo de Excel existente
    #workbook = Workbook()
    libro = load_workbook(ruta)
    
    # Seleccionar la hoja activa (por defecto es la primera hoja)
    hoja2 = libro.active
    hoja2.title = "hoja2"
    
    while arduino.is_open:
        # Leer datos del puerto serial
        data = arduino.readline().decode().strip()
        
        # Mostrar datos en label
        etiqueta1.config(text=data)
        
        # Obtener la siguiente fila vacía
        next_row = hoja2.max_row + 1
        
        # Guardar datos en archivo de Excel
        hoja2.cell(row=next_row, column=1).value = data
        
        # Guardar el archivo de Excel
        libro.save(ruta)
           

# Función para iniciar hilo
def start_thread():
    thread = threading.Thread(target=read_and_save)
    thread.daemon = True
    thread.start()
#----------------------------------------
def pressButton1():
    arduino.write(("1\n").encode())
    etiqueta2.config(text="Salida Encendida", fg="blue")

def pressButton2():
    arduino.write(("2\n").encode())
    etiqueta2.config(text="Salida Apagada", fg="red")

#----------------------------------------
# Crear ventana
root = Tk()
root.title("Lectura de Sensor")
root.geometry("400x200")
root.protocol("WM_DELETE_WINDOW", cerrar_ventana)

root.columnconfigure(0, minsize=200)

# Crear label para mostrar datos
etiqueta1 = Label(root, text="Esperando datos...", font=("Arial", 16))
etiqueta1.grid(row=0, column=0, padx=10, pady=10)

# Botón para iniciar lectura de sensor
boton_start = Button(root, text="Iniciar", font=("Arial", 12), command=start_thread)
boton_start.grid(row=1, column=0, padx=10, pady=10)

# Botón para detener lectura de sensor
boton_stop = Button(root, text="Detener", font=("Arial", 12), command=cerrar_ventana)
boton_stop.grid(row=2, column=0, padx=10, pady=10)

#------------------------------
button1 = Button(root, text=" ON", command=pressButton1, padx=10, pady=10)
button1.grid(row=0, column=1)

button2 = Button(root, text="OFF", command=pressButton2, padx=10, pady=10)
button2.grid(row=1, column=1)

etiqueta2 = Label(root, text="Salida Apagada", padx = 20, fg="red",  font=("Arial", 15))
etiqueta2.grid(row=0, column=2)
#------------------------------
root.minsize(450, 200)

root.mainloop()
""" 
  TAREA: Proyecto final : foto Dalia
  
"""