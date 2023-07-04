from openpyxl import*
import os

carpeta = r"C:\Users\xedkr\Desktop\clase5b"
ruta = r"C:\Users\xedkr\Desktop\clase5b\file2.xlsx"

#crear directorio o carpeta si no existe
if not os.path.exists(carpeta):
    os.makedirs(carpeta)

#crear archivo excel si no existe
if not os.path.exists(ruta):
    libro = Workbook()
    libro.save(ruta)

#cargar libro/archivo excel
libro = load_workbook(ruta)

#obtener la hoja/ crear la hoja
hoja1 = libro.active
hoja1.title = "Hoja_1"

#añadir datos a la hoja
hoja1["A1"] = "Nombre"
hoja1["B1"] = "Edad"
hoja1["A2"] = "Juan"
hoja1["B2"] = 25
hoja1["A3"] = "Maria"
hoja1["B3"] = 30

#Guardar datos
libro.save(ruta)
import tkinter as tk
import serial
import os
import threading
from openpyxl import *

carpeta = r"C:\Users\xedkr\Desktop\clase5b"
ruta = r"C:\Users\xedkr\Desktop\clase5b\sensorExl.xlsx"
puerto = "COM5"

# Crear directorio si no existe
if not os.path.exists(carpeta):
    os.makedirs(carpeta)

# Crear archivo de Excel
if not os.path.exists(ruta):
    libro = Workbook()
    libro.save(ruta)

# Crear puerto serial
try:
    arduino = serial.Serial(puerto, 9600)
except:
    print("No se estableció la conexión")


#para cerrar ventana
def cerrar_ventana():
    arduino.close()
    root.destroy()


# Función para leer y guardar datos en archivo de Excel
def read_and_save():
    # Cargar el archivo de Excel existente
    #workbook = Workbook()
    libro = load_workbook(ruta)
    
    # Seleccionar la hoja activa (por defecto es la primera hoja)
    hoja1 = libro.active
    hoja1.title = "hoja1"
    
    while arduino.is_open:        
        # Leer datos del puerto serial
        data = arduino.readline().decode().strip()
        
        # Mostrar datos en label
        etiqueta1.config(text=data)
        
        # Obtener la siguiente fila vacía
        next_row = hoja1.max_row + 1
        
        # Guardar datos en archivo de Excel
        hoja1.cell(row=next_row, column=1).value = data
        
        # Guardar el archivo de Excel
        libro.save(ruta)

# Función para iniciar hilo
def iniciar_hilo():
    hilo = threading.Thread(target=read_and_save)
    hilo.daemon = True
    hilo.start()

# Crear ventana
root = tk.Tk()
root.title("Lectura de Sensor")
root.geometry("400x200")
root.protocol("WM_DELETE_WINDOW", cerrar_ventana)

# Crear label para mostrar datos
etiqueta1 = tk.Label(root, text="Esperando dato.", font=("Arial", 16))
etiqueta1.grid(row=0, column=0, padx=10, pady=10)

# Botón para iniciar lectura de sensor
boton_start = tk.Button(root, text="Iniciar", font=("Arial", 12), command=iniciar_hilo)
boton_start.grid(row=1, column=0, padx=10, pady=10)

# Botón para detener lectura de sensor
boton_stop = tk.Button(root, text="Detener", font=("Arial", 12), command=cerrar_ventana)
boton_stop.grid(row=2, column=0, padx=10, pady=10)


root.mainloop()



""" 
  Ejercicio de tarea
  


"""